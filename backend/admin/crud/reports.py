import os
from datetime import datetime, timedelta
from typing import Optional

import httpx
import openpyxl
import schemas.reports as schemas
from config import settings
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


class ServiceClient:
    def __init__(self):
        self.db_client = httpx.AsyncClient(
            base_url=settings.DATABASE_SERVICE_URL, timeout=10.0
        )

    async def close(self):
        await self.db_client.aclose()


service_client = ServiceClient()


async def get_sales_summary(
    start_date: Optional[datetime] = None, end_date: Optional[datetime] = None
) -> schemas.SalesSummaryResponse:
    """Get sales summary from database API"""
    try:
        # Default to last 30 days if no dates provided
        if not end_date:
            end_date = datetime.utcnow()
        if not start_date:
            start_date = end_date - timedelta(days=30)

        # Build query params
        params = {}
        if start_date:
            params["start_date"] = start_date.isoformat()
        if end_date:
            params["end_date"] = end_date.isoformat()

        # Get data from database API
        summary_resp = await service_client.db_client.get(
            "/reports/sales/summary", params=params
        )
        top_products_resp = await service_client.db_client.get(
            "/reports/sales/top-products", params=params
        )
        by_day_resp = await service_client.db_client.get(
            "/reports/sales/by-day", params=params
        )
        by_hour_resp = await service_client.db_client.get(
            "/reports/sales/by-hour", params=params
        )

        if summary_resp.status_code != 200:
            raise Exception("Failed to fetch sales summary")

        summary = summary_resp.json()
        top_products = (
            top_products_resp.json() if top_products_resp.status_code == 200 else []
        )
        by_day = by_day_resp.json() if by_day_resp.status_code == 200 else []
        by_hour = by_hour_resp.json() if by_hour_resp.status_code == 200 else []

        return schemas.SalesSummaryResponse(
            total_sales=summary.get("total_sales", 0),
            total_orders=summary.get("total_orders", 0),
            total_items=summary.get("total_items", summary.get("items_sold", 0)),
            average_order_value=summary.get("average_order_value", 0),
            top_products=[
                schemas.TopProduct(
                    product_name=p["product_name"],
                    quantity=p.get("quantity", p.get("quantity_sold", 0)),
                    revenue=p["revenue"],
                )
                for p in top_products
            ],
            sales_by_day=[
                schemas.SalesByDay(
                    date=d["date"], total=d.get("total", d.get("sales", 0))
                )
                for d in by_day
            ],
            sales_by_hour=[
                schemas.SalesByHour(
                    hour=h["hour"], total=h.get("total", h.get("sales", 0))
                )
                for h in by_hour
            ],
        )

    except Exception as e:
        print(f"Error getting sales summary: {e}")
        return schemas.SalesSummaryResponse(
            total_sales=0,
            total_orders=0,
            total_items=0,
            average_order_value=0,
            top_products=[],
            sales_by_day=[],
            sales_by_hour=[],
        )


async def generate_excel_report(
    start_date: Optional[datetime] = None, end_date: Optional[datetime] = None
) -> Optional[str]:
    """Generate Excel report with sales data"""
    try:
        # Get sales data
        summary = await get_sales_summary(start_date, end_date)

        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Summary Sheet
        ws_summary = wb.create_sheet("Summary")

        # Header
        ws_summary["A1"] = "Sales Report"
        ws_summary["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        ws_summary["A1"].fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        ws_summary.merge_cells("A1:D1")

        # Date range
        date_format = "%d.%m.%Y"
        ws_summary["A2"] = (
            f"Period: {start_date.strftime(date_format) if start_date else 'All'} - {end_date.strftime(date_format) if end_date else 'All'}"
        )
        ws_summary.merge_cells("A2:D2")

        # Metrics
        ws_summary["A4"] = "Metric"
        ws_summary["B4"] = "Value"
        ws_summary["A4"].font = Font(bold=True)
        ws_summary["B4"].font = Font(bold=True)

        metrics = [
            ("Total Sales", f"{summary.total_sales:,.0f} so'm"),
            ("Total Orders", summary.total_orders),
            ("Total Items Sold", summary.total_items),
            ("Average Order Value", f"{summary.average_order_value:,.0f} so'm"),
        ]

        row = 5
        for metric, value in metrics:
            ws_summary[f"A{row}"] = metric
            ws_summary[f"B{row}"] = value
            row += 1

        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 20

        # Top Products Sheet
        ws_products = wb.create_sheet("Top Products")
        ws_products["A1"] = "Product Name"
        ws_products["B1"] = "Quantity Sold"
        ws_products["C1"] = "Revenue (so'm)"

        for cell in ["A1", "B1", "C1"]:
            ws_products[cell].font = Font(bold=True, color="FFFFFF")
            ws_products[cell].fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )

        row = 2
        for product in summary.top_products:
            ws_products[f"A{row}"] = product.product_name
            ws_products[f"B{row}"] = product.quantity
            ws_products[f"C{row}"] = product.revenue
            row += 1

        ws_products.column_dimensions["A"].width = 30
        ws_products.column_dimensions["B"].width = 15
        ws_products.column_dimensions["C"].width = 20

        # Daily Sales Sheet
        ws_daily = wb.create_sheet("Daily Sales")
        ws_daily["A1"] = "Date"
        ws_daily["B1"] = "Sales (so'm)"

        for cell in ["A1", "B1"]:
            ws_daily[cell].font = Font(bold=True, color="FFFFFF")
            ws_daily[cell].fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )

        row = 2
        for day in summary.sales_by_day:
            ws_daily[f"A{row}"] = day.date
            ws_daily[f"B{row}"] = day.total
            row += 1

        ws_daily.column_dimensions["A"].width = 15
        ws_daily.column_dimensions["B"].width = 20

        # Add chart
        if len(summary.sales_by_day) > 0:
            chart = LineChart()
            chart.title = "Daily Sales Trend"
            chart.y_axis.title = "Sales (so'm)"
            chart.x_axis.title = "Date"

            data = Reference(
                ws_daily, min_col=2, min_row=1, max_row=len(summary.sales_by_day) + 1
            )
            cats = Reference(
                ws_daily, min_col=1, min_row=2, max_row=len(summary.sales_by_day) + 1
            )

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            ws_daily.add_chart(chart, "D2")

        # Save file
        filename = f"/app/reports/sales_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        os.makedirs("/app/reports", exist_ok=True)
        wb.save(filename)

        return filename

    except Exception as e:
        print(f"Error generating Excel report: {e}")
        return None


async def get_inventory_report() -> schemas.InventoryReportResponse:
    """Get inventory status report from database API"""
    try:
        response = await service_client.db_client.get("/reports/inventory")

        if response.status_code != 200:
            raise Exception("Failed to fetch inventory report")

        data = response.json()

        return schemas.InventoryReportResponse(
            total_products=data.get("total_products", 0),
            low_stock_count=data.get("low_stock_count", 0),
            out_of_stock_count=data.get("out_of_stock_count", 0),
            total_value=data.get("total_value", 0),
            products=[
                schemas.InventoryItem(
                    product_name=p["product_name"],
                    quantity=p["quantity"],
                    price=p["price"],
                    value=p["value"],
                    status=p["status"],
                )
                for p in data.get("products", [])
            ],
        )

    except Exception as e:
        print(f"Error getting inventory report: {e}")
        return schemas.InventoryReportResponse(
            total_products=0,
            low_stock_count=0,
            out_of_stock_count=0,
            total_value=0,
            products=[],
        )
